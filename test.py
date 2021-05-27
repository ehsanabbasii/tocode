# -*- coding: utf-8 -*-


from openpyxl import Workbook
import matplotlib 
from openpyxl import load_workbook
import pandas as pd
import get_data 
import portfuireader
from time import sleep


workb = load_workbook(r"E:\py\userdata.xlsx")
#wb = load_workbook("E:\py\excel\salam.xlsx")

works=workb.active
print(workb.sheetnames)



every_thing = {}


def all_ids():#anjam chand taqir
    global every_thing
    for row in works.iter_rows(values_only=True):
        temp_list=[]
        if row[0]!= 'username':
            
            temp_list.append(row[1])
            every_thing[str(row[0])]=temp_list
        
            get_data.get_mofid()
            get_data.login(str(row[0]),row[1])
            
            sleep(10)
            get_data.download_excel()
            
            sleep(2)
            
            get_data.close()
            every_thing[row[0]].append(portfuireader.modify_rows(str(row[0])))
        
            sleep(2)
        
    


def search_data(name):
    havers=[]
    not_havers=[]
    global every_thing
    for user in every_thing.keys():
        print(user)
        
        for i in every_thing[user]:
            if type(i)== list:
                for k in i :
                    
           
                    if (name in k) == True :
                        print("***********",user)
                        havers.append(user)
              
            
    print("thease accounts have the mentioned stock :{:}".format(havers))
    
    for user in every_thing:
        if user not in havers:
            not_havers.append(user)
            
            
    print("thease accounts don't have the mentioned stock :{:}".format(not_havers))                 
       
        
        

all_ids()

name=input('what stock do you want to check? ')
search_data(name)


print(every_thing)
with open("E:\py\data.py","w",encoding="utf-8") as f: #in write mode
    f.write("data={}".format(every_thing))
    f.write("\ndeveloper_name={}".format("'ehsan'"))










'''

get_data.get_mofid()
get_data.login(a,b)
get_data.download_excel()

'''

