# -*- coding: utf-8 -*-
"""
Created on Wed Feb  3 17:56:44 2021

@author: EHSAN
"""

from openpyxl import Workbook
import matplotlib 
from openpyxl import load_workbook
import pandas as pd
import os




    
    
 
    
    
    

#name("num1")
#
#    wb = load_workbook("E:\py\excel\پورتفوی.xlsx")
##wb = load_workbook("E:\py\excel\salam.xlsx")
#
#    ws=wb.active
#    print(wb.sheetnames)
#
#




def rial_remover(a):
    num='0123456789'
    b=''
    for i in a:
        if i in num:
            b+=i
        
    return int(b)


def modify_rows(username):#anjam chand taqir
    
    def re_name(a):
    
        os.rename("پورتفوی.xlsx",a+".xlsx")
    
    
    wb = load_workbook(r"E:\py\پورتفوی.xlsx")
#wb = load_workbook("E:\py\excel\salam.xlsx")

    ws=wb.active
    print(wb.sheetnames)

    
    
    
    all_stock=[]
    
    
    
    
    for row in ws.iter_rows(values_only=True):
        
       
        
        if row[3]==None:
            pass
          #  ws.delete_rows(idx=c)
        
        
        elif row[0]  =="تعداد نماد"  :
            
            print("stock num {:}".format(row[1]))
            print("all poses {:}".format(rial_remover(row[3])))
        
        else :
            
               if row not in all_stock:
                all_stock.append(row)
                print(row)
            
        
            
    
    re_name(username)           
        
#modify_rows()

#print(all_stock[1:])


    
    with open("E:\py\data{:}.py".format(username),"w",encoding="utf-8") as f: #in write mode
        f.write("all_stk ={}".format(all_stock[1:]))
        
        f.write("\nname={}".format("'ehsan'"))


    ws.delete_rows(idx=0)
    df=pd.DataFrame(all_stock[1:],columns=all_stock[0])
    df.drop_duplicates(keep=False, inplace=True)
#df.values
    df.to_csv('E:\py\portfuy.csv')
    df
    return all_stock[1:]




#with open("E:\py\excel\sample.txt",encoding="utf-8") as f: #in read mode, not in write mode, careful ecoding is important!
 #   rd=f.readlines()
'''from sample import all_stk,name
print(all_stk)
print(name)  
'''




